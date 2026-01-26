import openpyxl

from .IParser import IParser
from ..logger import log


class XlsxParserWithHeader(IParser):

    def __init__(self, filename: str, sheet_number: int, record_start_with: str):
        self.file = openpyxl.load_workbook(filename, read_only=True)
        sheet_name = self.file.sheetnames[sheet_number]
        self.sheet = self.file[sheet_name]

        if record_start_with:
            # As file reader starts from index 0
            self.record_start_with = int(record_start_with)
        else:
            self.record_start_with = 0

        self.header = self.__get_header()

    def __get_header(self):
        headers = {}
        no_of_columns = self.sheet.max_column
        no_of_rows = self.sheet.max_row

        # check for end of file
        if self.record_start_with > no_of_rows:
            log.error("File probably dont contain any Record")
            return -1

        row = self.sheet[self.record_start_with]
        self.record_start_with = self.record_start_with + 1

        if not row:
            return -1
        else:
            for col in range(0, no_of_columns):
                if row[col].value:
                    headers[col] = row[col].value.strip()

        return headers

    def get_next_data(self):
        """
        Read and get single record in a dictionary in a xlsx file.
        Will return -1 when it reaches end
        """
        key_value = {}
        no_of_rows = self.sheet.max_row

        # check for end of file
        if self.record_start_with > no_of_rows:
            return -1

        row = self.sheet[self.record_start_with]

        # Unexpected end of file or data_header not set
        if not row or not self.header:
            return -1
        # if it reaches empty line in the middle of the file.
        if not any(cell.value for cell in row[:self.sheet.max_column]):
            return -1

        self.__add_values_with_data_header_moving_startpointer(key_value, row, self.header)
        return key_value

    def __add_values_with_data_header_moving_startpointer(self, key_value, row, data_header):
        # Read and assign column from data_header
        for column_num, column_name in data_header.items():
            key_value[column_name] = row[column_num].value
        self.record_start_with = self.record_start_with + 1

    def close(self):
        self.file.close()
